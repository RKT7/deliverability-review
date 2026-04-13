import openpyxl
from io import BytesIO

# ── Dimension order (must match PPT template) ──
DIM_ORDER = [
    "Scope Clarity & Client Readiness",
    "Key Dependencies & Third Parties",
    "Previous Work & Accelerators",
    "Transition to Operations & Sustainability",
    "Ease of New Deliverables Development"
]

def parse_excel(file):
    """Parse the filled Excel assessment file.
    
    Returns:
        qs: list of question dicts with keys: id, dim, question, response, team, rag, justification, action
        ov: deal overview text from 01_Deal_Overview
        risks: list of risk dicts with keys: risk, mit, owner
    """
    wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True, keep_vba=False)
    
    # ── Parse 02_Assessment ──
    qs = []
    if '02_Assessment' in wb.sheetnames:
        ws = wb['02_Assessment']
        for row in range(2, 23):  # rows 2-22, 21 questions
            q = {}
            q['id'] = ws.cell(row=row, column=1).value or (row - 1)  # Col A
            q['dim'] = str(ws.cell(row=row, column=2).value or '')    # Col B - dimension name
            q['question'] = str(ws.cell(row=row, column=4).value or '')  # Col D
            q['response'] = str(ws.cell(row=row, column=6).value or '')  # Col F
            q['team'] = str(ws.cell(row=row, column=8).value or '')      # Col H
            q['rag'] = str(ws.cell(row=row, column=9).value or '')       # Col I - formula result
            q['justification'] = str(ws.cell(row=row, column=10).value or '')  # Col J
            q['action'] = str(ws.cell(row=row, column=11).value or '')         # Col K
            
            if q['question'] and q['question'] != 'None':
                qs.append(q)
    
    # ── Parse 01_Deal_Overview ──
    ov = ''
    if '01_Deal_Overview' in wb.sheetnames:
        ws = wb['01_Deal_Overview']
        ov = str(ws.cell(row=3, column=1).value or '')
    
    # ── Parse 03_Risks ──
    risks = []
    if '03_Risks' in wb.sheetnames:
        ws = wb['03_Risks']
        for row in range(2, 11):  # rows 2-10
            r = {}
            r['risk'] = str(ws.cell(row=row, column=1).value or '')   # Col A
            r['mit'] = str(ws.cell(row=row, column=2).value or '')    # Col B
            r['owner'] = str(ws.cell(row=row, column=3).value or '')  # Col C
            if r['risk'] and r['risk'] != 'None':
                risks.append(r)
    
    wb.close()
    return qs, ov, risks


def calc_rags(qs):
    """Calculate RAG scores per dimension and overall.
    
    Rules:
    - Any RED in dimension = dimension RED
    - Any AMBER (no RED) = dimension AMBER
    - All GREEN = dimension GREEN
    - Any RED dimension = Overall RED
    - Any AMBER dimension = Overall AMBER
    - All GREEN = Overall GREEN
    """
    dim_rags = {d: 'GREEN' for d in DIM_ORDER}
    
    for q in qs:
        rag = (q.get('rag') or '').strip().lower()
        dim = (q.get('dim') or '').strip()
        
        if not rag or rag == 'n/a' or rag == 'none':
            continue
        
        # Match dimension using partial string matching (lowercased)
        matched_dim = None
        for d in DIM_ORDER:
            if d.lower() in dim.lower() or dim.lower() in d.lower():
                matched_dim = d
                break
        
        if not matched_dim:
            continue
        
        if rag == 'red':
            dim_rags[matched_dim] = 'RED'
        elif rag == 'amber' and dim_rags[matched_dim] != 'RED':
            dim_rags[matched_dim] = 'AMBER'
    
    vals = list(dim_rags.values())
    if 'RED' in vals:
        overall = 'RED'
    elif 'AMBER' in vals:
        overall = 'AMBER'
    else:
        overall = 'GREEN'
    
    return {'dimRags': dim_rags, 'overall': overall}
